import os
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from openpyxl.styles import borders
from openpyxl.worksheet.dimensions import ColumnDimension
import warnings

warnings.simplefilter("ignore")

# Streamlit App
st.title("PTML Site Data Base Management")

# Input and Output File Paths
#input_path = st.text_input("Input Excel File Path 📂", "C:/Users/UWX161178/S_Shah_Sb/Cells_DB_Mid_Dec_2024.xlsx")
input_path = st.file_uploader("Upload Excel File 📂", type=["xlsx"])
output_path = st.text_input("Output Excel File Path 📤", "C:/Users/UWX161178/S_Shah_Sb/PTML_Cell_List.xlsx")

# Fixed Column Order
def_columns = {
    "2G": ['Tech region', 'Site ID', 'Site Type', 'Cell ID simple', 'Current Hgt', 'Beam Width', 'Current Azimuth',
           'Current E-Tilt', 'New MSC ID', 'New BSC', 'LAC', 'CGI', 'City Name', 'Province', 'District', 'Tehsil',
           'Sector Name', 'Covered Area', 'BSIC', 'BCCH ARFCN', 'Long', 'Degree', 'Min', 'Sec', 'Latitude',
           'GSM Antenna', 'DCS Antenna', 'DB Antenna', 'TRIB Antenna', 'Total Antenna Count', 'PMO Status'],
    
    "3G": ['Tech region', '2G Site ID', '3G Site ID', 'CL Site Tech', 'Freq. Band', 'Cell ID simple', 'PSC',
           'RNC ID', 'LAC', 'CGI', '3G Site Name', 'Current Hgt', 'Current Azimuth', 'Current E-Tilt', 'City',
           'Province', 'District', 'Tehsil', 'Longitude', 'Latitude', 'Site Type', 'Frequency DOWNLINK',
           'Frequency UPLINK', 'Horizontal BW', 'Vertical BW', 'Antenna Type', 'PMO Status'],
    
    "4G": ['Tech region', '4G Site ID', 'Cell No.', '2G Site ID', '3G Site ID', 'eNodeB ID', '4G spectrum BW',
           'Cell Freq. Band', 'CL Site Tech', 'ECI', 'ECGI', 'TAC', '4G Site Name', 'Current Hgt',
           'Current Azimuth', 'Current E-Tilt', 'Latitude', 'Longitude', 'Site Type', 'City', 'Province',
           'District', 'Tehsil', 'New Antenna Type', 'PMO Status']
}

@st.cache_data
def load_sheets(file_path):
    """Load sheet names from the Excel file."""
    try:
        return pd.ExcelFile(file_path).sheet_names
    except Exception as e:
        st.error(f"Error loading Excel file: {e}")
        return []

# Load available sheets
sheet_names = load_sheets(input_path)

# Default sheet selection
selected_sheets = {
    "2G": st.selectbox("Select 2G Sheet Name 📄", sheet_names, index=sheet_names.index("2G") if "2G" in sheet_names else 0),
    "3G": st.selectbox("Select 3G Sheet Name 📄", sheet_names, index=sheet_names.index("3G") if "3G" in sheet_names else 0),
    "4G": st.selectbox("Select 4G Sheet Name 📄", sheet_names, index=sheet_names.index("4G") if "4G" in sheet_names else 0)
}

# Let user modify required columns (without column order input)
selected_columns = {
    "2G": st.multiselect("Select Required Columns for 2G 📝", def_columns["2G"], default=def_columns["2G"]),
    "3G": st.multiselect("Select Required Columns for 3G 📝", def_columns["3G"], default=def_columns["3G"]),
    "4G": st.multiselect("Select Required Columns for 4G 📝", def_columns["4G"], default=def_columns["4G"]),
}

# Let user select PMO Status values
pmo_status_options = ["CL", "NCL", "NA", "NCL-relocation", "Planned"]
selected_pmo_status = st.multiselect("Select PMO Status values 🛠️", pmo_status_options, default=pmo_status_options)

# Load and filter data
def load_and_filter_data(sheet_name, selected_cols, pmo_status_values):
    """Load data from the selected sheet, enforce selected column order, and filter rows based on PMO Status."""
    try:
        df = pd.read_excel(input_path, sheet_name=sheet_name, dtype={'ECGI': str})
        df = df[[col for col in selected_cols if col in df.columns]]
        df = df[df['PMO Status'].isin(pmo_status_values)].reset_index(drop=True).drop(columns=['PMO Status'])
        return df
    except Exception as e:
        st.error(f"Error processing sheet {sheet_name}: {e}")
        return pd.DataFrame()

# Format Excel file
def format_excel(file_path):
# Set Tab Color (All the Tabs)
    """Apply formatting to the Excel file."""
    wb = openpyxl.load_workbook(file_path)
    colors = ["00B0F0", "0000FF", "ADD8E6", "87CEFA"]
    
    for i, ws in enumerate(wb):
        ws.sheet_properties.tabColor = colors[i % len(colors)]

# Font, Alignment and Border (All the Sheets)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    
    style = NamedStyle(name="styled_cell")
    style.font = Font(name='Calibri Light', size=8)
    style.alignment = Alignment(horizontal='center', vertical='center')
    style.border = border
    wb.add_named_style(style)
    
    wb.calculation.calcMode = 'manual'
    
    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                cell.style = "styled_cell"
    
    wb.calculation.calcMode = 'auto'

# WrapText of header and Formatting (All the sheets)    
    for ws in wb:
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=11, name='Calibri Light')

# Set Filter on the Header (All the sheet)    
    for ws in wb:
        # Get the first row
        first_row = ws[1]
        # Apply the filter on the first row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(first_row))}1"


# Set Column Width (All the sheets)
    for ws in wb.worksheets:
        # Iterate over all columns in the sheet
        for column in ws.columns:
            # Get the current width of the column
            current_width = ws.column_dimensions[column[0].column_letter].width
            # Get the maximum width of the cells in the column
            length = max(len(str(cell.value)) for cell in column)
            # Set the width of the column to fit the maximum width, if it's greater than the current width
            if length > current_width:
                ws.column_dimensions[column[0].column_letter].width = length
    
# Insert a New Sheet (as First Sheet)    
    ws_title = wb.create_sheet("Title Page", 0)
    # Merge Specific Row and Columns
    ws_title.merge_cells(start_row=12, start_column=5, end_row=18, end_column=17)
    # Fill the Merge Cells
    ws_title.cell(row=12, column=5).value = 'PTML Network Site DataBase'

# Formatting Tital Page Report    
    # Access the first row starting from row 3
    first_row1 = list(ws_title.rows)[11]
    # Iterate through the cells in the first row starting from column E
    for cell in first_row1[4:]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.fill = openpyxl.styles.PatternFill(start_color="CF0A2C", end_color="CF0A2C", fill_type = "solid")
        font = openpyxl.styles.Font(color="FFFFFF",bold=True,size=60,name='Calibri Light')
        cell.font = font

# Hide the gridlines    
    ws_title.sheet_view.showGridLines = False
# Hide the headings
    ws_title.sheet_view.showRowColHeaders = False

# Hyper Link For Title Page
    # loop through all sheets in the workbook and insert the hyperlink to each sheet
    row = 22
    for ws in wb:
        if ws.title != "Title Page":
            hyperlink_cell = ws_title.cell(row=row, column=5)
            hyperlink_cell.value = ws.title
            hyperlink_cell.hyperlink = "#'{}'!A1".format(ws.title)
            hyperlink_cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
            #hyperlink_cell.border = border
            hyperlink_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            # set the height of the cell
            ws_title.row_dimensions[row].height = 15
            # set the colum width of column 5
            ws_title.column_dimensions[get_column_letter(5)].width = 30
            row += 1

# Hyper Link For Sub Pages    
    # Loop through all sheets in the workbook and insert the hyperlink to each sheet
    for i, ws in enumerate(wb.worksheets):
        # Check if the sheet is not the Title Page
        if ws.title != "Title Page":
            # Add hyperlink to cell in the last column+2 of the sheet
            hyperlink_cell = ws.cell(row=2, column=ws.max_column+2)
            hyperlink_cell.value = "Back to Table of Contents"
            hyperlink_cell.hyperlink = "#'{}'!E{}".format("Title Page", 22)
            hyperlink_cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
            hyperlink_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            # Set border on the hyperlink column
            for row in ws.iter_rows(min_row=2, max_row=4, min_col=hyperlink_cell.column, max_col=hyperlink_cell.column):
                for cell in row:
                    cell.border = border
            # Set width of the hyperlink column
            col_letter = openpyxl.utils.get_column_letter(hyperlink_cell.column)
            ws.column_dimensions[col_letter].width = 25

            # Add hyperlink to cell in the last column+2 of the sheet for next sheet
            if i < len(wb.worksheets)-1:
                next_hyperlink_cell = ws.cell(row=3, column=ws.max_column)
                next_hyperlink_cell.value = "Next Sheet"
                next_hyperlink_cell.hyperlink = "#'{}'!A1".format(wb.worksheets[i+1].title)
                next_hyperlink_cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
                next_hyperlink_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Add hyperlink to cell in the last column+2 of the sheet for previous sheet
            if i > 0:
                prev_hyperlink_cell = ws.cell(row=4, column=ws.max_column)
                prev_hyperlink_cell.value = "Previous Sheet"
                prev_hyperlink_cell.hyperlink = "#'{}'!A1".format(wb.worksheets[i-1].title)
                prev_hyperlink_cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
                prev_hyperlink_cell.alignment = openpyxl.styles.Alignment(horizontal='center')

    wb.save(file_path)
    return file_path

# Process and export data
if st.button("Process and Export Data 🚀"):
    with st.spinner("Processing data... ⏳"):
        df_2g = load_and_filter_data(selected_sheets["2G"], selected_columns["2G"], selected_pmo_status)
        df_3g = load_and_filter_data(selected_sheets["3G"], selected_columns["3G"], selected_pmo_status)
        df_4g = load_and_filter_data(selected_sheets["4G"], selected_columns["4G"], selected_pmo_status)

        with pd.ExcelWriter(output_path) as writer:
            df_2g.to_excel(writer, sheet_name="2G_Cells", index=False)
            df_3g.to_excel(writer, sheet_name="3G_Cells", index=False)
            df_4g.to_excel(writer, sheet_name="4G_Cells", index=False)

        #st.success(f"Filtered data saved to {output_path} ✅")

        # Format the file
        formatted_file = format_excel(output_path)
        st.success(f"Formatted file saved: {formatted_file} 🎨")
