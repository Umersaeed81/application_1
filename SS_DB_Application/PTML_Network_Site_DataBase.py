# import required libraries
import os
import sys
import openpyxl
import warnings
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl.styles import borders
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
#-----------------------------------------------------------------------------------------------#
# Set the page title and favicon (logo in browser tab)
st.set_page_config(
    page_title="PTML Network Site DataBase",  # Title of the web page
    page_icon="D:/Advance_Data_Sets/PTML_DB/Huawei.jpg",  # Path to the image file you want to use as favicon
)
#-----------------------------------------------------------------------------------------------#
# Create two columns
col1, col2 = st.columns([2, 1])  # Adjust column width ratio as needed

# Left column: Markdown with personal details
with col1:
    st.markdown("""
        # [Umer Saeed](https://www.linkedin.com/in/engumersaeed/)
        Sr. RF Planning & Optimization Engineer  
        BSc Telecommunications Engineering, School of Engineering  
        MS Data Science, School of Business and Economics  
        **University of Management & Technology**  
        **Mobile:**     +923018412180  
        **Email:**  umersaeed81@hotmail.com  
        **Address:** Dream Gardens, Defence Road, Lahore  
    """)

# Right column: Logo image centered
with col2:
    # Use a placeholder to center the image within the column
    col2.markdown("<div style='display: flex; justify-content: center;'>", unsafe_allow_html=True)
    col2.image("D:/Advance_Data_Sets/PTML_DB/Huawei.jpg", width=100)
    col2.markdown("</div>", unsafe_allow_html=True)
#-----------------------------------------------------------------------------------------------#
# Replace with your given date in YYYY-MM-DD format
given_date_str = "2026-01-03"  # Example date
given_date = datetime.strptime(given_date_str, "%Y-%m-%d")
# Get today's date
today_date = datetime.today()

# Compare dates
if given_date < today_date:
    print("OK.")
    sys.exit(0)  # Exit the script
else:
    print("NoK.")
#-----------------------------------------------------------------------------------------------#     
# Streamlit App
st.markdown("<h1 style='color: maroon;'>PTML Network Site DataBase</h1>", unsafe_allow_html=True)
#-----------------------------------------------------------------------------------------------# 
st.markdown("<h3 style='color: maroon;'>Input and Output File Path</h3>", unsafe_allow_html=True)
#-----------------------------------------------------------------------------------------------# 
# Input File Path
input_path = st.text_input("**Input Excel File Path 📂**", "D:/Advance_Data_Sets/PTML_DB/Cells_DB_Mid_Dec_2024.xlsx")

# Validate file path
if not input_path.lower().endswith('.xlsx'):
    st.error("Only .xlsx files are allowed! Please provide a valid file.")
elif not os.path.exists(input_path):
    st.error(f"The file does not exist: {input_path}")
else:
    try:
        # Read all sheets at once (Optimized step)
        xls_data = pd.read_excel(input_path, sheet_name=None, dtype=str)
        
        # Extract sheet names
        sheet_names = list(xls_data.keys())
        num_sheets = len(sheet_names)

        if num_sheets < 3:
            st.error(f"The Excel file must contain at least 3 sheets! Found only {num_sheets}.")
        else:
            missing_column_sheets = []

            # Check for 'PMO Status' column in each sheet
            for sheet, df in xls_data.items():
                df.columns = df.columns.str.strip()  # Clean column names
                if 'PMO Status' not in df.columns:
                    missing_column_sheets.append(sheet)
            
            if missing_column_sheets:
                st.error(f"The following sheets are missing the 'PMO Status' column: {', '.join(missing_column_sheets)}")
            else:
                st.success(f"Valid file with {num_sheets} sheets, all containing 'PMO Status'.")

    except Exception as e:
        st.error(f"Error reading the Excel file: {e}")

#-----------------------------------------------------------------------------------------------#
# Output File Path Generation
today_date = datetime.today().strftime("%d%m%Y")
base_path = f"D:/Advance_Data_Sets/PTML_DB/PTML_Cell_List_{today_date}.xlsx"

# Ensure directory exists
if not os.path.exists(os.path.dirname(base_path)):
    st.error(f"Error: The specified directory {os.path.dirname(base_path)} does not exist.")
else:
    def get_unique_filename(base_path):
        """Generate a unique filename if the file already exists."""
        counter = 1
        new_filename = base_path
        while os.path.exists(new_filename):
            new_filename = base_path.replace(".xlsx", f"_{counter}.xlsx")
            counter += 1
        return new_filename

    output_path = get_unique_filename(base_path)
    user_output_path = st.text_input("**Output Excel File Path 📤**", output_path)

    if os.path.exists(user_output_path):
        st.warning("The file already exists, a unique name will be generated.")
    else:
        st.success(f"File path is valid: {user_output_path}")

#-----------------------------------------------------------------------------------------------#
# Fixed Column Order
def_columns = {
    "2G": ['Tech region', 'Site ID', 'Site Type', 'Cell ID simple', 'Current Hgt', 'Beam Width', 'Current Azimuth',
           'Current E-Tilt', 'New MSC ID', 'New BSC', 'LAC', 'CGI', 'City Name', 'Province', 'District', 'Tehsil',
           'Sector Name', 'Covered Area', 'BSIC', 'BCCH ARFCN', 'Long', 'Degree', 'Min', 'Sec', 'Latitude',
           'GSM Antenna', 'DCS Antenna', 'DB Antenna', 'TRIB Antenna', 'Total Antenna Count', 'PMO Status'],
    
    "3G": ['Tech region', '2G Site ID', '3G Site ID', 'CL Site Tech', 'Freq. Band', 'Cell ID simple', 'PSC',
           'RNC ID', 'LAC', 'CGI', '3G Site Name', 'Current Hgt', 'Current Azimuth', 'Current E-Tilt', 'City',
           'Province', 'District', 'Tehsil', 'Longitude', 'Latitude', 'Site Type', 'Frequency DOWNLINK',
           'Frequency UPLINK', 'Horizontal BW', 'Vertical BW', 'Antenna Type', 'PMO Status'],
    
    "4G": ['Tech region', '4G Site ID', 'Cell No.', '2G Site ID', '3G Site ID', 'eNodeB ID', '4G spectrum BW',
           'Cell Freq. Band', 'CL Site Tech', 'ECI', 'ECGI', 'TAC', '4G Site Name', 'Current Hgt',
           'Current Azimuth', 'Current E-Tilt', 'Latitude', 'Longitude', 'Site Type', 'City', 'Province',
           'District', 'Tehsil', 'New Antenna Type', 'PMO Status']
}
#-----------------------------------------------------------------------------------------------#
# Sheet Selection
st.markdown("<h3 style='color: maroon;'>Select Sheet Name</h3>", unsafe_allow_html=True)
selected_sheets = {
    "2G": st.selectbox("**Select 2G Sheet Name 📄**", sheet_names, index=sheet_names.index("2G") if "2G" in sheet_names else 0),
    "3G": st.selectbox("**Select 3G Sheet Name 📄**", sheet_names, index=sheet_names.index("3G") if "3G" in sheet_names else 0),
    "4G": st.selectbox("**Select 4G Sheet Name 📄**", sheet_names, index=sheet_names.index("4G") if "4G" in sheet_names else 0)
}
#-----------------------------------------------------------------------------------------------#
# Column Selection
st.markdown("<h3 style='color: maroon;'>Select the Required Columns</h3>", unsafe_allow_html=True)
col1, col2, col3 = st.columns(3)

selected_columns = {
    "2G": col1.multiselect("**Select Required Columns for 2G 📝**", def_columns["2G"], default=def_columns["2G"]),
    "3G": col2.multiselect("**Select Required Columns for 3G 📝**", def_columns["3G"], default=def_columns["3G"]),
    "4G": col3.multiselect("**Select Required Columns for 4G 📝**", def_columns["4G"], default=def_columns["4G"])
}
#-----------------------------------------------------------------------------------------------#
# Validate PMO Status Selection
missing_pmo = [tech for tech, cols in selected_columns.items() if "PMO Status" not in cols]
if missing_pmo:
    st.warning(f"⚠️ 'PMO Status' must be selected for {', '.join(missing_pmo)}")

# Collect PMO Status values only from the selected sheets
pmo_status_set = set()

for tech, sheet in selected_sheets.items():
    if sheet in xls_data:  # Ensure the sheet exists
        df = xls_data[sheet]
        df.columns = df.columns.str.strip()  # Clean column names
        
        if "PMO Status" in df.columns:
            df["PMO Status"] = df["PMO Status"].fillna("NA").str.strip()
            pmo_status_set.update(df["PMO Status"].unique())  # Collect unique PMO values from selected sheets

# PMO Status Selection
st.markdown("<h3 style='color: maroon;'>Select Required PMO Status</h3>", unsafe_allow_html=True)
default_selected = {"CL", "NCL"}

st.write("**Select PMO Status values 🛠️**")
selected_pmo_status = [status for status in sorted(pmo_status_set) if st.checkbox(status, value=(status in default_selected))]

#-----------------------------------------------------------------------------------------------#

# Load and filter data
def load_and_filter_data(sheet_name, selected_cols, pmo_status_values):
    """Load data from the selected sheet, enforce selected column order, and filter rows based on PMO Status."""
    try:
        df = pd.read_excel(input_path, sheet_name=sheet_name, dtype={'ECGI': str})
        df = df[[col for col in selected_cols if col in df.columns]]
        df = df[df['PMO Status'].isin(pmo_status_values)].reset_index(drop=True).drop(columns=['PMO Status'])
        return df
    except Exception as e:
        st.error(f"Error processing sheet {sheet_name}: {e}")
        return pd.DataFrame()

#-----------------------------------------------------------------------------------------------#        
# Format Excel file
def format_excel(file_path):
# Set Tab Color (All the Tabs)
    """Apply formatting to the Excel file."""
    wb = openpyxl.load_workbook(file_path)
    steps = 10  # Total formatting steps
    progress_bar = st.progress(0)  # Initialize progress bar
    step = 0
    progress_bar.progress(step / steps)
    colors = ["00B0F0", "0000FF", "ADD8E6", "87CEFA"] 
    for i, ws in enumerate(wb):
        ws.sheet_properties.tabColor = colors[i % len(colors)]
    step += 1
    progress_bar.progress(step / steps)
# #-----------------------------------------------------------------------------------------------#
# Font, Alignment and Border (All the Sheets)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    
    style = NamedStyle(name="styled_cell")
    style.font = Font(name='Calibri Light', size=8)
    style.alignment = Alignment(horizontal='center', vertical='center')
    style.border = border
    wb.add_named_style(style)
    
    wb.calculation.calcMode = 'manual'
    
    for ws in wb:
        for row in ws.iter_rows():
            for cell in row:
                cell.style = "styled_cell"
    
    wb.calculation.calcMode = 'auto'
    step += 1
    progress_bar.progress(step / steps)
# #-----------------------------------------------------------------------------------------------#
# WrapText of header and Formatting (All the sheets)    
    for ws in wb:
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=11, name='Calibri Light')
    step += 1
    progress_bar.progress(step / steps)
# #-----------------------------------------------------------------------------------------------#
# Set Filter on the Header (All the sheet)    
    for ws in wb:
        # Get the first row
        first_row = ws[1]
        # Apply the filter on the first row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(first_row))}1"
    step += 1
    progress_bar.progress(step / steps)
# #-----------------------------------------------------------------------------------------------#
    for ws in wb:
        # Get the first row
        first_row = ws[1]
        # Freeze the top row
        ws.freeze_panes = "A2"
    step += 1
    progress_bar.progress(step / steps)
#---------------------------------
# Set Column Width (All the sheets)
    for ws in wb.worksheets:
        # Iterate over all columns in the sheet
        for column in ws.columns:
            # Get the current width of the column
            current_width = ws.column_dimensions[column[0].column_letter].width
            # Get the maximum width of the cells in the column
            length = max(len(str(cell.value)) for cell in column)
            # Set the width of the column to fit the maximum width, if it's greater than the current width
            if length > current_width:
                ws.column_dimensions[column[0].column_letter].width = length
    step += 1
    progress_bar.progress(step / steps)
# #-----------------------------------------------------------------------------------------------#    
# # Insert a New Sheet (as First Sheet)    
    ws_title = wb.create_sheet("Title Page", 0)
    # Merge Specific Row and Columns
    ws_title.merge_cells(start_row=12, start_column=5, end_row=18, end_column=17)
    # Fill the Merge Cells
    ws_title.cell(row=12, column=5).value = 'PTML Network Site DataBase'
    step += 1
    progress_bar.progress(step / steps)
# #-----------------------------------------------------------------------------------------------#
# # Formatting Tital Page Report    
    # Access the first row starting from row 3
    first_row1 = list(ws_title.rows)[11]
    # Iterate through the cells in the first row starting from column E
    for cell in first_row1[4:]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.fill = openpyxl.styles.PatternFill(start_color="CF0A2C", end_color="CF0A2C", fill_type = "solid")
        font = openpyxl.styles.Font(color="FFFFFF",bold=True,size=60,name='Calibri Light')
        cell.font = font
    step += 1
    progress_bar.progress(step / steps)
# #-----------------------------------------------------------------------------------------------#
# # Hide the gridlines    
    ws_title.sheet_view.showGridLines = False
# Hide the headings
    ws_title.sheet_view.showRowColHeaders = False
# #-----------------------------------------------------------------------------------------------#
# # Hyper Link For Title Page
    # loop through all sheets in the workbook and insert the hyperlink to each sheet
    row = 22
    for ws in wb:
        if ws.title != "Title Page":
            hyperlink_cell = ws_title.cell(row=row, column=5)
            hyperlink_cell.value = ws.title
            hyperlink_cell.hyperlink = "#'{}'!A1".format(ws.title)
            hyperlink_cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
            #hyperlink_cell.border = border
            hyperlink_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            # set the height of the cell
            ws_title.row_dimensions[row].height = 15
            # set the colum width of column 5
            ws_title.column_dimensions[get_column_letter(5)].width = 30
            row += 1
    step += 1
    progress_bar.progress(step / steps)
# #-----------------------------------------------------------------------------------------------#
# Hyper Link For Sub Pages    
    # Loop through all sheets in the workbook and insert the hyperlink to each sheet
    for i, ws in enumerate(wb.worksheets):
        # Check if the sheet is not the Title Page
        if ws.title != "Title Page":
            # Add hyperlink to cell in the last column+2 of the sheet
            hyperlink_cell = ws.cell(row=2, column=ws.max_column+2)
            hyperlink_cell.value = "Back to Table of Contents"
            hyperlink_cell.hyperlink = "#'{}'!E{}".format("Title Page", 22)
            hyperlink_cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
            hyperlink_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            # Set border on the hyperlink column
            for row in ws.iter_rows(min_row=2, max_row=4, min_col=hyperlink_cell.column, max_col=hyperlink_cell.column):
                for cell in row:
                    cell.border = border
            # Set width of the hyperlink column
            col_letter = openpyxl.utils.get_column_letter(hyperlink_cell.column)
            ws.column_dimensions[col_letter].width = 25

            # Add hyperlink to cell in the last column+2 of the sheet for next sheet
            if i < len(wb.worksheets)-1:
                next_hyperlink_cell = ws.cell(row=3, column=ws.max_column)
                next_hyperlink_cell.value = "Next Sheet"
                next_hyperlink_cell.hyperlink = "#'{}'!A1".format(wb.worksheets[i+1].title)
                next_hyperlink_cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
                next_hyperlink_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Add hyperlink to cell in the last column+2 of the sheet for previous sheet
            if i > 0:
                prev_hyperlink_cell = ws.cell(row=4, column=ws.max_column)
                prev_hyperlink_cell.value = "Previous Sheet"
                prev_hyperlink_cell.hyperlink = "#'{}'!A1".format(wb.worksheets[i-1].title)
                prev_hyperlink_cell.font = openpyxl.styles.Font(color="0000FF", underline="single")
                prev_hyperlink_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    step += 1
    progress_bar.progress(step / steps)           
# #-----------------------------------------------------------------------------------------------#
# ## Inset Image on the title page    
    # inset the Huawei logo
    img = Image('D:/Advance_Data_Sets/PTML_DB/Huawei.jpg')
    img.width = 7 * 15
    img.height = 7 * 15
    ws_title.add_image(img,'E3')
# #-----------------------------------------------------------------------------------------------#
#     # inset the PTCL logo
    img1 = Image('D:/Advance_Data_Sets/PTML_DB/PTCL.png')
    ws_title.add_image(img1,'M3')
# #-----------------------------------------------------------------------------------------------#
    wb.save(file_path)
    return file_path
# #-----------------------------------------------------------------------------------------------#
st.markdown("<h3 style='color: maroon;'>Export Output</h3>", unsafe_allow_html=True)

# Process and export data
if st.button("Process and Export Data 🚀"):
    with st.spinner("Processing data... ⏳"):
        df_2g = load_and_filter_data(selected_sheets["2G"], selected_columns["2G"], selected_pmo_status)
        df_3g = load_and_filter_data(selected_sheets["3G"], selected_columns["3G"], selected_pmo_status)
        df_4g = load_and_filter_data(selected_sheets["4G"], selected_columns["4G"], selected_pmo_status)

        with pd.ExcelWriter(output_path) as writer:
            df_2g.to_excel(writer, sheet_name="2G_Cells", index=False)
            df_3g.to_excel(writer, sheet_name="3G_Cells", index=False)
            df_4g.to_excel(writer, sheet_name="4G_Cells", index=False)

        #st.success(f"Filtered data saved to {output_path} ✅")

        # Format the file
        formatted_file = format_excel(output_path)
        st.success(f"Formatted file saved: {formatted_file} 🎨")
#-----------------------------------------------------------------------------------------------#